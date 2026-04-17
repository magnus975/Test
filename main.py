import json
import os
import sqlite3
import io
import logging
import requests
from datetime import datetime, timezone

from flask import Flask, g, jsonify, request, send_file, send_from_directory
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(APP_DIR, "inventory.db")
ALERTS_PATH = os.path.join(APP_DIR, "pending_alerts.json")

# For Render persistence: use /var/data if available (Render Disk), otherwise APP_DIR.
# Attach a Render Disk at /var/data and set env INVENTORY_DB_DIR=/var/data for persistence.
_db_dir = os.environ.get("INVENTORY_DB_DIR")
if _db_dir and os.path.isdir(_db_dir):
    DB_PATH = os.path.join(_db_dir, "inventory.db")
    ALERTS_PATH = os.path.join(_db_dir, "pending_alerts.json")

# Wingman API configuration for sending emails
WINGMAN_API_BASE = os.environ.get("WINGMAN_API_BASE", "")
WINGMAN_API_TOKEN = os.environ.get("WINGMAN_API_TOKEN", "")

app = Flask(__name__, static_folder="static")
with app.app_context():
	init_db()
logger = logging.getLogger(__name__)


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            min_quantity INTEGER NOT NULL DEFAULT 0,
            current_quantity INTEGER NOT NULL DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
        """
    )
    conn.execute(
        "INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)",
        ("alert_email", "magnus@ostin.no"),
    )
    conn.commit()
    conn.close()


def get_setting(key, default=None):
    """Read a setting from the database (usable outside Flask request context)."""
    conn = sqlite3.connect(DB_PATH)
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    conn.close()
    return row[0] if row else default


# --- Email Sending ---

def _build_alert_email_html(low_stock_items):
    """Build a formatted HTML email body for low-stock alerts."""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    rows_html = ""
    for item in low_stock_items:
        deficit = item["min_quantity"] - item["current_quantity"]
        rows_html += (
            f'<tr style="border-bottom:1px solid #e5e7eb;">'
            f'<td style="padding:10px 14px;font-weight:500;">{item["product_name"]}</td>'
            f'<td style="padding:10px 14px;text-align:center;">{item["current_quantity"]}</td>'
            f'<td style="padding:10px 14px;text-align:center;">{item["min_quantity"]}</td>'
            f'<td style="padding:10px 14px;text-align:center;color:#dc2626;font-weight:600;">-{deficit}</td>'
            f'</tr>'
        )

    html = f"""
    <div style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;max-width:600px;margin:0 auto;">
      <div style="background:#0284c7;color:white;padding:20px 24px;border-radius:12px 12px 0 0;">
        <h2 style="margin:0;font-size:20px;">⚠️ Low Stock Alert</h2>
        <p style="margin:6px 0 0;opacity:0.85;font-size:14px;">{len(low_stock_items)} product{"s" if len(low_stock_items) != 1 else ""} below minimum — {now}</p>
      </div>
      <div style="border:1px solid #e5e7eb;border-top:none;border-radius:0 0 12px 12px;overflow:hidden;">
        <table style="width:100%;border-collapse:collapse;font-size:14px;">
          <thead>
            <tr style="background:#f9fafb;">
              <th style="padding:10px 14px;text-align:left;font-weight:600;color:#374151;">Product</th>
              <th style="padding:10px 14px;text-align:center;font-weight:600;color:#374151;">Current</th>
              <th style="padding:10px 14px;text-align:center;font-weight:600;color:#374151;">Minimum</th>
              <th style="padding:10px 14px;text-align:center;font-weight:600;color:#374151;">Deficit</th>
            </tr>
          </thead>
          <tbody>
            {rows_html}
          </tbody>
        </table>
      </div>
      <p style="margin-top:16px;font-size:12px;color:#6b7280;">Sent by Inventory Manager</p>
    </div>
    """
    return html


def _build_alert_email_plain(low_stock_items):
    """Build plain-text fallback for low-stock alerts."""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    lines = [
        f"LOW STOCK ALERT — {now}",
        f"{len(low_stock_items)} product(s) below minimum stock level:",
        "",
    ]
    for item in low_stock_items:
        deficit = item["min_quantity"] - item["current_quantity"]
        lines.append(
            f"  • {item['product_name']}: {item['current_quantity']}/{item['min_quantity']} (need {deficit} more)"
        )
    lines.append("")
    lines.append("— Inventory Manager")
    return "\n".join(lines)


def send_alert_email(to_email, low_stock_items):
    """Send a low-stock alert email via the Gmail integration.

    Returns a dict with 'success' (bool) and 'message' (str).
    The function writes a pending_email.json file that Material (the Wingman
    agent) picks up and sends via GMAIL_SEND_EMAIL.  If WINGMAN_API_BASE and
    WINGMAN_API_TOKEN are configured, it also attempts to call the API directly.
    """
    if not to_email:
        return {"success": False, "message": "No alert email configured"}
    if not low_stock_items:
        return {"success": False, "message": "No low-stock items to report"}

    count = len(low_stock_items)
    subject = f"⚠️ Low Stock Alert: {count} product{'s' if count != 1 else ''} below minimum"
    body_html = _build_alert_email_html(low_stock_items)
    body_plain = _build_alert_email_plain(low_stock_items)

    # Always write the pending email file so the Wingman agent can pick it up
    email_payload = {
        "to_email": to_email,
        "subject": subject,
        "body_html": body_html,
        "body_plain": body_plain,
        "low_stock_items": low_stock_items,
        "generated_at": datetime.now(timezone.utc).isoformat() + "Z",
        "status": "pending",
    }

    pending_email_path = os.path.join(os.path.dirname(ALERTS_PATH), "pending_email.json")
    with open(pending_email_path, "w") as f:
        json.dump(email_payload, f, indent=2)

    # Attempt direct API call if credentials are configured
    if WINGMAN_API_BASE and WINGMAN_API_TOKEN:
        try:
            resp = requests.post(
                f"{WINGMAN_API_BASE}/api/v1/execute",
                headers={
                    "Authorization": f"Bearer {WINGMAN_API_TOKEN}",
                    "Content-Type": "application/json",
                },
                json={
                    "tool_name": "GMAIL_SEND_EMAIL",
                    "arguments": json.dumps({
                        "recipient_email": to_email,
                        "subject": subject,
                        "body": body_html,
                        "is_html": True,
                    }),
                },
                timeout=30,
            )
            if resp.status_code == 200:
                email_payload["status"] = "sent"
                with open(pending_email_path, "w") as f:
                    json.dump(email_payload, f, indent=2)
                return {"success": True, "message": f"Email sent to {to_email}"}
            else:
                logger.warning("Wingman API returned %s: %s", resp.status_code, resp.text)
                email_payload["status"] = "api_error"
                email_payload["api_error"] = resp.text[:500]
                with open(pending_email_path, "w") as f:
                    json.dump(email_payload, f, indent=2)
        except Exception as e:
            logger.warning("Failed to call Wingman API: %s", e)
            email_payload["status"] = "api_unreachable"
            with open(pending_email_path, "w") as f:
                json.dump(email_payload, f, indent=2)

    # If API wasn't configured or failed, the pending_email.json is still written
    # for the Wingman agent to pick up and send.
    if email_payload["status"] == "pending":
        return {
            "success": True,
            "message": f"Alert email queued for {to_email} (pending_email.json written)",
        }
    elif email_payload["status"] == "sent":
        return {"success": True, "message": f"Email sent to {to_email}"}
    else:
        return {
            "success": True,
            "message": f"Alert email queued for {to_email} (API call failed, pending_email.json written for agent pickup)",
        }


# --- Routes ---

@app.route("/")
def index():
    return send_from_directory("static", "index.html")


@app.route("/api/products", methods=["GET"])
def list_products():
    db = get_db()
    rows = db.execute(
        "SELECT id, name, min_quantity, current_quantity FROM products ORDER BY name"
    ).fetchall()
    products = [dict(r) for r in rows]
    return jsonify(products)


@app.route("/api/products", methods=["POST"])
def add_product():
    data = request.get_json(force=True)
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Product name is required"}), 400

    min_qty = data.get("min_quantity")
    if min_qty is None or not str(min_qty).lstrip("-").isdigit():
        return jsonify({"error": "Minimum quantity must be a number"}), 400
    min_qty = int(min_qty)
    if min_qty < 0:
        return jsonify({"error": "Minimum quantity cannot be negative"}), 400

    current_qty = int(data.get("current_quantity", 0))
    if current_qty < 0:
        return jsonify({"error": "Current quantity cannot be negative"}), 400

    db = get_db()
    try:
        db.execute(
            "INSERT INTO products (name, min_quantity, current_quantity) VALUES (?, ?, ?)",
            (name, min_qty, current_qty),
        )
        db.commit()
    except sqlite3.IntegrityError:
        return jsonify({"error": f"Product '{name}' already exists"}), 409

    row = db.execute(
        "SELECT id, name, min_quantity, current_quantity FROM products WHERE name = ?",
        (name,),
    ).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/products/<int:product_id>", methods=["PATCH"])
def update_product(product_id):
    data = request.get_json(force=True)
    db = get_db()

    row = db.execute("SELECT id FROM products WHERE id = ?", (product_id,)).fetchone()
    if not row:
        return jsonify({"error": "Product not found"}), 404

    updates = []
    params = []

    if "current_quantity" in data:
        qty = data["current_quantity"]
        if not str(qty).lstrip("-").isdigit():
            return jsonify({"error": "Current quantity must be a number"}), 400
        qty = int(qty)
        if qty < 0:
            return jsonify({"error": "Current quantity cannot be negative"}), 400
        updates.append("current_quantity = ?")
        params.append(qty)

    if "min_quantity" in data:
        mq = data["min_quantity"]
        if not str(mq).lstrip("-").isdigit():
            return jsonify({"error": "Minimum quantity must be a number"}), 400
        mq = int(mq)
        if mq < 0:
            return jsonify({"error": "Minimum quantity cannot be negative"}), 400
        updates.append("min_quantity = ?")
        params.append(mq)

    if "name" in data:
        name = (data["name"] or "").strip()
        if not name:
            return jsonify({"error": "Product name is required"}), 400
        updates.append("name = ?")
        params.append(name)

    if not updates:
        return jsonify({"error": "No fields to update"}), 400

    params.append(product_id)
    db.execute(f"UPDATE products SET {', '.join(updates)} WHERE id = ?", params)
    db.commit()

    row = db.execute(
        "SELECT id, name, min_quantity, current_quantity FROM products WHERE id = ?",
        (product_id,),
    ).fetchone()
    return jsonify(dict(row))


@app.route("/api/products/<int:product_id>", methods=["DELETE"])
def delete_product(product_id):
    db = get_db()
    row = db.execute("SELECT id FROM products WHERE id = ?", (product_id,)).fetchone()
    if not row:
        return jsonify({"error": "Product not found"}), 404
    db.execute("DELETE FROM products WHERE id = ?", (product_id,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/check-alerts", methods=["POST"])
def check_alerts():
    db = get_db()
    rows = db.execute(
        """
        SELECT name, current_quantity, min_quantity
        FROM products
        WHERE current_quantity < min_quantity
        ORDER BY name
        """
    ).fetchall()

    alerts = [
        {
            "product_name": r["name"],
            "current_quantity": r["current_quantity"],
            "min_quantity": r["min_quantity"],
        }
        for r in rows
    ]

    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat() + "Z",
        "low_stock_count": len(alerts),
        "alerts": alerts,
    }

    with open(ALERTS_PATH, "w") as f:
        json.dump(payload, f, indent=2)

    # Send the alert email if there are low-stock items
    email_result = {"success": False, "message": "No low-stock items"}
    if alerts:
        alert_email = get_setting("alert_email")
        if alert_email:
            email_result = send_alert_email(alert_email, alerts)
        else:
            email_result = {"success": False, "message": "No alert email configured in settings"}
    else:
        email_result = {"success": True, "message": "All products are fully stocked — no email needed"}

    payload["email_status"] = email_result

    return jsonify(payload)


@app.route("/api/pending-alerts", methods=["GET"])
def get_pending_alerts():
    if not os.path.exists(ALERTS_PATH):
        return jsonify({"low_stock_count": 0, "alerts": []})
    with open(ALERTS_PATH) as f:
        return jsonify(json.load(f))


@app.route("/api/pending-email", methods=["GET"])
def get_pending_email():
    """Return the pending email payload (for agent pickup)."""
    pending_path = os.path.join(os.path.dirname(ALERTS_PATH), "pending_email.json")
    if not os.path.exists(pending_path):
        return jsonify({"status": "none", "message": "No pending email"})
    with open(pending_path) as f:
        return jsonify(json.load(f))


@app.route("/api/export-xlsx", methods=["GET"])
def export_xlsx():
    db = get_db()
    rows = db.execute(
        "SELECT name, min_quantity, current_quantity FROM products ORDER BY name"
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # -- Styles --
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    low_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    low_font = Font(color="B91C1C")
    ok_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    ok_font = Font(color="15803D")
    center_align = Alignment(horizontal="center")

    # -- Header row --
    headers = ["Product Name", "Min Quantity", "Current Quantity", "Status"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # -- Data rows --
    for row_idx, r in enumerate(rows, 2):
        is_low = r["current_quantity"] < r["min_quantity"]
        status = "⚠ Low Stock" if is_low else "✓ OK"

        ws.cell(row=row_idx, column=1, value=r["name"]).border = thin_border
        c_min = ws.cell(row=row_idx, column=2, value=r["min_quantity"])
        c_min.alignment = center_align
        c_min.border = thin_border
        c_cur = ws.cell(row=row_idx, column=3, value=r["current_quantity"])
        c_cur.alignment = center_align
        c_cur.border = thin_border
        c_status = ws.cell(row=row_idx, column=4, value=status)
        c_status.alignment = center_align
        c_status.border = thin_border

        if is_low:
            c_status.fill = low_fill
            c_status.font = low_font
        else:
            c_status.fill = ok_fill
            c_status.font = ok_font

    # -- Column widths --
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"inventory_{timestamp}.xlsx"

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/settings", methods=["GET"])
def get_settings():
    db = get_db()
    rows = db.execute("SELECT key, value FROM settings").fetchall()
    settings = {r["key"]: r["value"] for r in rows}
    return jsonify(settings)


@app.route("/api/settings", methods=["POST"])
def update_settings():
    data = request.get_json(force=True)
    if not data or not isinstance(data, dict):
        return jsonify({"error": "Request body must be a JSON object"}), 400

    db = get_db()
    for key, value in data.items():
        if not isinstance(key, str) or not isinstance(value, str):
            return jsonify({"error": f"Both key and value must be strings"}), 400
        db.execute(
            "INSERT INTO settings (key, value) VALUES (?, ?) "
            "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
            (key, value),
        )
    db.commit()

    rows = db.execute("SELECT key, value FROM settings").fetchall()
    settings = {r["key"]: r["value"] for r in rows}
    return jsonify(settings)


if __name__ == "__main__":
    init_db()
    print(f"✅ Inventory Management App running at http://0.0.0.0:5000")
    app.run(host="0.0.0.0", port=5000, debug=False)
