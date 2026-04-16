import io
import json
import os
import sqlite3
import tempfile

import pytest
from openpyxl import load_workbook

# Patch DB/alert paths before importing app
_tmp_dir = tempfile.mkdtemp()
os.environ["TESTING"] = "1"

import main

main.DB_PATH = os.path.join(_tmp_dir, "test_inventory.db")
main.ALERTS_PATH = os.path.join(_tmp_dir, "test_alerts.json")


@pytest.fixture(autouse=True)
def fresh_db():
    """Re-init DB for each test."""
    if os.path.exists(main.DB_PATH):
        os.remove(main.DB_PATH)
    main.init_db()
    yield
    if os.path.exists(main.DB_PATH):
        os.remove(main.DB_PATH)


@pytest.fixture
def client():
    main.app.config["TESTING"] = True
    with main.app.test_client() as c:
        yield c


def _add_product(client, name, min_qty, cur_qty):
    return client.post(
        "/api/products",
        data=json.dumps({"name": name, "min_quantity": min_qty, "current_quantity": cur_qty}),
        content_type="application/json",
    )


class TestExportXlsx:
    def test_export_empty_inventory(self, client):
        """Export with no products returns a valid xlsx with only headers."""
        resp = client.get("/api/export-xlsx")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.content_type

        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb.active
        assert ws.title == "Inventory"
        assert ws.cell(1, 1).value == "Product Name"
        assert ws.cell(1, 2).value == "Min Quantity"
        assert ws.cell(1, 3).value == "Current Quantity"
        assert ws.cell(1, 4).value == "Status"
        # No data rows
        assert ws.cell(2, 1).value is None

    def test_export_with_products(self, client):
        """Export includes all products with correct data."""
        _add_product(client, "Widget A", 10, 20)
        _add_product(client, "Widget B", 50, 5)
        _add_product(client, "Widget C", 10, 10)

        resp = client.get("/api/export-xlsx")
        assert resp.status_code == 200

        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb.active

        # 3 products + 1 header = 4 rows
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 3

        # Sorted by name: Widget A, Widget B, Widget C
        names = [row[0] for row in data_rows]
        assert names == ["Widget A", "Widget B", "Widget C"]

    def test_export_low_stock_status(self, client):
        """Low-stock products get '⚠ Low Stock' status; others get '✓ OK'."""
        _add_product(client, "Stocked Item", 5, 50)
        _add_product(client, "Low Item", 100, 3)

        resp = client.get("/api/export-xlsx")
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb.active

        rows = {r[0]: r for r in ws.iter_rows(min_row=2, values_only=True)}
        assert rows["Low Item"][3] == "⚠ Low Stock"
        assert rows["Stocked Item"][3] == "✓ OK"

    def test_export_quantities_correct(self, client):
        """Min and current quantities match what was added."""
        _add_product(client, "Gadget", 25, 12)

        resp = client.get("/api/export-xlsx")
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb.active

        assert ws.cell(2, 1).value == "Gadget"
        assert ws.cell(2, 2).value == 25
        assert ws.cell(2, 3).value == 12

    def test_export_content_disposition(self, client):
        """Response has attachment content-disposition with .xlsx filename."""
        resp = client.get("/api/export-xlsx")
        cd = resp.headers.get("Content-Disposition", "")
        assert "attachment" in cd
        assert "inventory_" in cd
        assert ".xlsx" in cd

    def test_export_styling(self, client):
        """Headers have the brand-blue fill and bold white font."""
        _add_product(client, "Test", 5, 10)

        resp = client.get("/api/export-xlsx")
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb.active

        header_cell = ws.cell(1, 1)
        assert header_cell.font.bold is True
        # Header fill should be brand blue 0284C7
        assert header_cell.fill.start_color.rgb == "FF0284C7" or header_cell.fill.start_color.rgb == "000284C7" or "0284C7" in str(header_cell.fill.start_color.rgb)

    def test_export_after_quantity_update(self, client):
        """Export reflects updated quantities."""
        _add_product(client, "Dynamic Item", 10, 50)

        # Update quantity
        products = json.loads(client.get("/api/products").data)
        pid = products[0]["id"]
        client.patch(
            f"/api/products/{pid}",
            data=json.dumps({"current_quantity": 3}),
            content_type="application/json",
        )

        resp = client.get("/api/export-xlsx")
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb.active

        assert ws.cell(2, 3).value == 3
        assert ws.cell(2, 4).value == "⚠ Low Stock"
